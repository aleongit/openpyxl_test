from openpyxl import load_workbook

# Simple usage_______________________________________________________________
# Fold (outline)_________________________________________

# file
file = 'usage.xlsx'
sheetname = 'Fold'

# open file
wb = load_workbook(filename = file)

# if sheet exists
if sheetname in wb.sheetnames:
    print(f'{sheetname} exists!!!')
    # select sheet
    sheet = wb[sheetname]

else:
    print(f'{sheetname} NO exists!!!')
    # new sheet at the end
    sheet = wb.create_sheet(title=sheetname)

# fold
sheet.column_dimensions.group('A','D', hidden=True)
sheet.row_dimensions.group(1,10, hidden=True)

# save
wb.save(filename = file)