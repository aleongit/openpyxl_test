from openpyxl import Workbook

# Optimised Modes_______________________________________________________________
# Write-only mode_________________________________________

# Here again, the regular openpyxl.worksheet.worksheet.
# Worksheet has been replaced by a faster alternative,
#   the openpyxl.worksheet._write_only.WriteOnlyWorksheet.
# When you want to dump large amounts of data make sure you have lxml installed.

wb = Workbook(write_only=True)
ws = wb.create_sheet()

# now we'll fill it with 100 rows x 200 columns

for irow in range(100):
    # ws.append(['%d' % i for i in range(200)])
    ws.append([i for i in range(200)])

# save the file
wb.save('big_file.xlsx') # doctest: +SKIP


# If you want to have cells with styles or comments then use a openpyxl.cell.WriteOnlyCell()

# from openpyxl import Workbook

wb = Workbook(write_only = True)
ws = wb.create_sheet()

from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

cell = WriteOnlyCell(ws, value="hello world")
cell.font = Font(name='Courier', size=36)
cell.comment = Comment(text="A comment", author="Author's Name")
ws.append([cell, 3.14, None])

wb.save('write_only_file.xlsx')