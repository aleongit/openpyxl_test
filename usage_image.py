from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Simple usage_______________________________________________________________
# Inserting an image_________________________________________

# file
file = 'usage.xlsx'
sheetname = 'Image'

# open file
wb = load_workbook(filename = file)

# if sheet exists
if sheetname in wb.sheetnames:
    print(f'{sheetname} exists!!!')
    # select sheet
    sheet = wb[sheetname]

else:
    print(f'{sheetname} NO exists!!!')
    # new sheet at the end
    sheet = wb.create_sheet(title=sheetname)

# ini
sheet['A1'] = 'You should see three logos below'

# create an image
img = Image('images/logo.png')

# add to worksheet and anchor next to cells
sheet.add_image(img, 'A1')

# save
wb.save(filename = file)