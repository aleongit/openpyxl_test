import datetime
from openpyxl import load_workbook

# Simple usage_______________________________________________________________
# Using number formats_________________________________________

# file
file = 'usage.xlsx'
sheetname = 'Number_Format'

# open file
wb = load_workbook(filename = file)

# if sheet exists
if sheetname in wb.sheetnames:
    print('%s exists' %(sheetname))
    print( sheetname + ' exists!')
    print(f'{sheetname} exists!!!')
    
    # select sheet
    sheet = wb[sheetname]

else:
    print('%s NO exists' %(sheetname))
    print( sheetname + ' NO exists!')
    print(f'{sheetname} NO exists!!!')
    
    # new sheet at the end
    sheet = wb.create_sheet(title=sheetname)

# set date using a Python datetime
sheet['A1'] = datetime.datetime(2010, 7, 21)
sheet['A1'].number_format
# 'yyyy-mm-dd h:mm:ss'

print(sheet['A1'].value)

# save
wb.save(filename = file)