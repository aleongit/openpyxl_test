import openpyxl
from openpyxl import load_workbook

# Defined Names_______________________________________________________________
# Sample use for ranges_________________________________________

# The specification has the following to say about defined names:

# Defined names are descriptive text that is used to represents a cell, range of cells, 
#   formula, or constant value.”
# This means they are very loosely defined. 
# They might contain a constant, a formula, a single cell reference, 
# a range of cells or multiple ranges of cells across different worksheets.
# Or all of the above.
# They are defined globally for a workbook and accessed from the defined_names attribute.

# open
file = 'defined_names.xlsx'
wb = load_workbook(filename = file)
ws = wb.worksheets[0]

# list defined names
print(wb.defined_names)
print()

poblacions = wb.defined_names['poblacions']
print(poblacions)
print()

# if this contains a range of cells then the destinations attribute is not None
dests = poblacions.destinations # returns a generator of (worksheet title, cell range) tuples
print(dests)

# get cells
cells = []
for title, coord in dests:
    print(title, coord)
    sheet = wb[title]
    cells.append(sheet[coord])

print(cells)
print()

# get values of defined range
for rang in cells:
    #print(rang)
    for row in rang:
        #print(row)
        for cell in row:
            #print(type(cell))
            #print(cell)
            print(cell.value, end = " ")
        print()

# Creating new named ranges_________________________________________

new_range = openpyxl.workbook.defined_name.DefinedName('newrange', attr_text='Hoja1!$E$1:$H$5')
wb.defined_names.append(new_range)

# create a local named range (only valid for a specific sheet)
sheetid = wb.sheetnames.index('Hoja1')
private_range = openpyxl.workbook.defined_name.DefinedName('privaterange', attr_text='Hoja1!$J$1:$J$5', localSheetId=sheetid)
wb.defined_names.append(private_range)
# this local range can't be retrieved from the global defined names
assert('privaterange' not in wb.defined_names)

# the scope has to be supplied to retrieve local ranges:
print(wb.defined_names.localnames(sheetid))
print(wb.defined_names.get('privaterange', sheetid).attr_text)

# save
wb.save(filename = file)