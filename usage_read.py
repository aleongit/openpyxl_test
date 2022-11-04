from openpyxl import load_workbook

# Simple usage_______________________________________________________________
# Read an existing workbook_________________________________________

wb = load_workbook(filename = 'usage.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
# 3

# -------------------------------------------------------------------------------------------
# Note !!!
# There are several flags that can be used in load_workbook.
# 
# . data_only controls whether cells with formulae have either the formula (default) 
# or the value stored the last time Excel read the sheet.
# 
# . keep_vba controls whether any Visual Basic elements are preserved or not (default). 
# If they are preserved they are still not editable.
# 
# Warning !!!
# openpyxl does currently not read all possible items in an Excel file 
# so images and charts will be lost from existing files if they are opened and saved with the same name.
# -------------------------------------------------------------------------------------------