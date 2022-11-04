from openpyxl import Workbook

# Create a workbook_______________________________________________________________

# There is no need to create a file on the filesystem to get started with openpyxl.
# Just import the Workbook class and start work:
wb = Workbook()

# A workbook is always created with at least one worksheet. 
# You can get it by using the Workbook.active property:
ws = wb.active

# -------------------------------------------------------------------------------------------
# Note !!!
# This is set to 0 by default. 
# Unless you modify its value, you will always get the first worksheet by using this method.
# -------------------------------------------------------------------------------------------

# You can create new worksheets using the Workbook.create_sheet() method:
ws1 = wb.create_sheet("End") # insert at the end (default)
# or
ws2 = wb.create_sheet("First", 0) # insert at first position
# or
ws3 = wb.create_sheet("Penultimate", -1) # insert at the penultimate position

# Sheets are given a name automatically when they are created.
# They are numbered in sequence (Sheet, Sheet1, Sheet2, …). 
# You can change this name at any time with the Worksheet.title property:
ws.title = "New"

# The background color of the tab holding this title is white by default. 
# You can change this providing an RRGGBB color code to the Worksheet.sheet_properties.tabColor attribute:

#ws.sheet_properties.tabColor = "1072BA"
ws.sheet_properties.tabColor = "E2FFA7"

# Once you gave a worksheet a name, you can get it as a key of the workbook:
ws = wb["New"]

# You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute
print(wb.sheetnames)
#['Sheet2', 'New Title', 'Sheet1']

# You can loop through worksheets
for sheet in wb:
    print(sheet.title)

# Set active sheet
# wb.active = wb['sheet_name']
wb.active = wb['New']
print(wb.active)

# You can create copies of worksheets within a single workbook with Workbook.copy_worksheet() method:
source = wb.active
target = wb.copy_worksheet(source)

# -------------------------------------------------------------------------------------------
# Note !!!
# Only cells (including values, styles, hyperlinks and comments) 
# and certain worksheet attribues (including dimensions, format and properties) are copied. 
# All other workbook / worksheet attributes are not copied - e.g. Images, Charts.
# You also cannot copy worksheets between workbooks. 
# You cannot copy a worksheet if the workbook is open in read-only or write-only mode.
# -------------------------------------------------------------------------------------------

# Playing with data_______________________________________________________________
# Accessing one cell_________________________________________

# Now we know how to get a worksheet, we can start modifying cells content. 
# Cells can be accessed directly as keys of the worksheet:
c = ws['A4']
print('cell', c)

# This will return the cell at A4, or create one if it does not exist yet.
# Values can be directly assigned:
ws['A1'] = 4
ws['A2'] = 'test'

# There is also the Worksheet.cell() method.
# This provides access to cells using row and column notation:
d = ws.cell(row=4, column=2, value=10)

# -------------------------------------------------------------------------------------------
# Note !!!
# When a worksheet is created in memory, it contains no cells.
# They are created when first accessed.
# 
# Warning !!!
# Because of this feature, scrolling through cells 
# instead of accessing them directly will create them all in memory, 
# even if you don’t assign them a value.
#
# Something like
#
# >>> for x in range(1,101):
# ...        for y in range(1,101):
# ...            ws.cell(row=x, column=y)
#
# will create 100x100 cells in memory, for nothing.
# -------------------------------------------------------------------------------------------


# Accessing many cells_________________________________________

# Ranges of cells can be accessed using slicing:
cell_range = ws['A1':'C2']
print(cell_range)

# Ranges of rows or columns can be obtained similarly:
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

print(colC)
print(col_range)
print(row10)
print(row_range)
print()

# You can also use the Worksheet.iter_rows() method:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
print()
# <Cell Sheet1.A1>
# <Cell Sheet1.B1>
# <Cell Sheet1.C1>
# <Cell Sheet1.A2>
# <Cell Sheet1.B2>
# <Cell Sheet1.C2>

# Likewise the Worksheet.iter_cols() method will return columns:
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
print()
# <Cell Sheet1.A1>
# <Cell Sheet1.A2>
# <Cell Sheet1.B1>
# <Cell Sheet1.B2>
# <Cell Sheet1.C1>
# <Cell Sheet1.C2>

# -------------------------------------------------------------------------------------------
# Note !!!
# For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.
# -------------------------------------------------------------------------------------------

# If you need to iterate through all the rows or columns of a file,
# you can instead use the Worksheet.rows property:
ws = wb.active
ws['C9'] = 'hello world'
rows = tuple(ws.rows)
print(rows)
print()

# or the Worksheet.columns property:
print(list(ws.columns))
print()

# -------------------------------------------------------------------------------------------
# Note !!!
# For performance reasons the Worksheet.columns property is not available in read-only mode.
# -------------------------------------------------------------------------------------------


# Values only_________________________________________

# If you just want the values from a worksheet you can use the Worksheet.values property.
# This iterates over all the rows in a worksheet but returns just the cell values:
for row in ws.values:
   for value in row:
     print(value)
print()

# Both Worksheet.iter_rows() and Worksheet.iter_cols()
# can take the values_only parameter to return just the cell’s value:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)
print()


# Data storage_______________________________________________________________

# Once we have a Cell, we can assign it a value:
c.value = 'hello, world'
print(c.value)
#'hello, world'

d.value = 3.14
print(d.value)
#3.14

# Saving to a file_________________________________________

# The simplest and safest way to save a workbook is by using the Workbook.save() method of the Workbook object:
# wb = Workbook()
# wb.save('balances.xlsx')

# -------------------------------------------------------------------------------------------
# Warning !!!
# This operation will overwrite existing files without warning.
#
# Note !!!
# The filename extension is not forced to be xlsx or xlsm,
# although you might have some trouble opening it directly with another application 
# if you don’t use an official extension.
# As OOXML files are basically ZIP files, you can also open it with your favourite ZIP archive manager.
# -------------------------------------------------------------------------------------------

# Saving as a stream_________________________________________

# If you want to save the file to a stream,
# e.g. when using a web application such as Pyramid, Flask or Django 
# then you can simply provide a NamedTemporaryFile():

'''
>>> from tempfile import NamedTemporaryFile
>>> from openpyxl import Workbook
>>> wb = Workbook()
>>> with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
'''

# You can specify the attribute template=True, to save a workbook as a template:

'''
>>> wb = load_workbook('document.xlsx')
>>> wb.template = True
>>> wb.save('document_template.xltx')
'''

# or set this attribute to False (default), to save as a document:

'''
>>> wb = load_workbook('document_template.xltx')
>>> wb.template = False
>>> wb.save('document.xlsx', as_template=False)
'''

# -------------------------------------------------------------------------------------------
# Warning !!!
# You should monitor the data attributes and document extensions 
# for saving documents in the document templates and vice versa,
# otherwise the result table engine can not open the document.
#
# Note !!!
# The following will fail:
#
# >>> wb = load_workbook('document.xlsx')
# >>> # Need to save with the extension *.xlsx
# >>> wb.save('new_document.xlsm')
# >>> # MS Excel can't open the document
# >>>
# >>> # or
# >>>
# >>> # Need specify attribute keep_vba=True
# >>> wb = load_workbook('document.xlsm')
# >>> wb.save('new_document.xlsm')
# >>> # MS Excel will not open the document
# >>>
# >>> # or
# >>>
# >>> wb = load_workbook('document.xltm', keep_vba=True)
# >>> # If we need a template document, then we must specify extension as *.xltm.
# >>> wb.save('new_document.xlsm')
# >>> # MS Excel will not open the document
# -------------------------------------------------------------------------------------------

# Save the file
wb.save("tutorial.xlsx")