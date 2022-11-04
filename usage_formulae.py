from openpyxl import load_workbook

# Simple usage_______________________________________________________________
# Using formulae_________________________________________

# file
file = 'usage.xlsx'
sheetname = 'Formulae'

# open file
wb = load_workbook(filename = file)

# if sheet exists
if sheetname in wb.sheetnames:
    print(f'{sheetname} exists!!!')
    # select sheet
    sheet = wb[sheetname]

else:
    print(f'{sheetname} NO exists!!!')
    # new sheet at the end
    sheet = wb.create_sheet(title=sheetname)

# add a simple formula
sheet["A1"] = "=SUM(1, 1)"
print(sheet['A1'].value)

# -------------------------------------------------------------------------------------------
# Warning !!!
# NB you must use the English name for a function and function arguments must be separated by commas 
# and not other punctuation such as semi-colons.
# -------------------------------------------------------------------------------------------

# openpyxl never evaluates formula but it is possible to check the name of a formula:

from openpyxl.utils import FORMULAE

print("HEX2DEC" in FORMULAE)
print("SUM" in FORMULAE)
#True

# If you’re trying to use a formula that isn’t known this could be because 
# you’re using a formula that was not included in the initial specification.
# Such formulae must be prefixed with _xlfn. to work.

# save
wb.save(filename = file)

