import datetime
from openpyxl import load_workbook

# Simple usage_______________________________________________________________
# Merge / Unmerge cells_________________________________________

# file
file = 'usage.xlsx'
sheetname = 'Merge'

# open file
wb = load_workbook(filename = file)

# if sheet exists
if sheetname in wb.sheetnames:
    print(f'{sheetname} exists!!!')
    # select sheet
    sheet = wb[sheetname]

else:
    print(f'{sheetname} NO exists!!!')
    # new sheet at the end
    sheet = wb.create_sheet(title=sheetname)

# When you merge cells all cells but the top-left one are removed from the worksheet. 
# To carry the border-information of the merged cell, 
# the boundary cells of the merged cell are created as MergeCells which always have the value None. 
# See Styling Merged Cells for information on formatting merged cells.

# ini
rang_merged = 'A2:D2'
merges = []

# list merged_cells.ranges
print(sheet.merged_cells.ranges)

for rang in sheet.merged_cells.ranges:
    print(rang)
    print(type(rang))
    merges.append(str(rang))    # to str
print(merges)

# si rang merged 
if rang_merged in merges:
  print("Oh no, the cell is merged!")
  sheet.unmerge_cells(rang_merged)
else:
  print("This cell is not merged.")
  sheet.merge_cells(rang_merged)

# or equivalently
# sheet.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
# sheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

# save
wb.save(filename = file)

